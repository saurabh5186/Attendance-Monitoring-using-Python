import face_recognition
import cv2
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook

now = datetime.now()
wb = Workbook()
ws = wb.active
d = now.strftime('%B %d (%a).xlsx') 

now = datetime.now()
dt = now.strftime("%H_%M_%S")
c = "D -["+dt+"].csv"   #creating file name with csv extention
#f = open(c, 'a')    #  creating excle file 

def save_n(name):
    # this function is use for saving the name of aendes 
    f = open(c, "a")
    f.write("\n"+name)
    f.close()
h = int(now.strftime("%H"))
ws1 = wb.create_sheet(str(h)+" to "+str(h+1),0)
ws1.append(["Name","Time"])

def save(name,t):
  if h!=int(now.strftime("%H")):
    x = now.strftime("%H")
    ws2 = wb.create_sheet(x,0)
  ws1.append([name,t])
 
#import file
# start of code
# Path of image attendence
path = 'ImgATTENDENCE'
images = []
Classname = []
myList = os.listdir(path)
# print(myList)
for cls in myList:
    curImg = cv2.imread(f'{path}/{cls}')
    images.append(curImg)
    Classname.append(os.path.splitext(cls)[0])

# print("className")
# print(Classname)
# encoding process
def findEncoding(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList

encodeListknown = findEncoding(images)
# print(len(encodeListknown))
cap = cv2.VideoCapture(1)

while True:
    success, img = cap.read(0)
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
    faceCureFrame = face_recognition.face_locations(imgS)
    endcodesCurFrame = face_recognition.face_encodings(imgS, faceCureFrame)

    for endcodeFace, faceLoc in zip(endcodesCurFrame, faceCureFrame):
        matches = face_recognition.compare_faces(encodeListknown, endcodeFace)
        faceDis = face_recognition.face_distance(encodeListknown, endcodeFace)
        # print(faceDis)
        matchIndex = np.argmin(faceDis)


        if matches[matchIndex]:
            name = Classname[matchIndex].upper()
            print(name)
            y1,x2,y2,x1 = faceLoc
            y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
            cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
            cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,255,0),cv2.FILLED)
            cv2.putText(img,name,(x1+6,y26),cv2.FONT_HERSHEY_COMPLEX,1,(255,255,255),2)
    t = now.strftime("%H:%M:%S")
    #save_n(name)
    save(name,t)
    wb.save(d)
    cv2.imshow('webcam', img)
    cv2.waitKey(1)